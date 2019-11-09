import re
import urllib.request
import requests
import sys,os
from openpyxl import Workbook
from openpyxl import load_workbook
f = open("timp.txt","r")
bpath = f.read()
f.close
base_clip_path = 'https://clips-media-assets2.twitch.tv/'


def retrieve_mp4_data(slug):
    cid = "cx2zppacbubmk2h9rolhr7gmdgwgoz"
    clip_info = requests.get(
        "https://api.twitch.tv/helix/clips?id=" + slug,
        headers={"Client-ID": cid}).json()
    thumb_url = clip_info['data'][0]['thumbnail_url']
    title = clip_info['data'][0]['title']
    slice_point = thumb_url.index("-preview-")
    mp4_url = thumb_url[:slice_point] + '.mp4'
    return mp4_url


def dl_progress(count, block_size, total_size):
    percent = int(count * block_size * 100 / total_size)
    sys.stdout.write("\r...%d%%" % percent)

workbook = load_workbook(bpath+'/Excel.xlsx')
ws = workbook.active
temp = 2
for clip in open(bpath + "/Slug.txt", 'r'):
    slug = clip.replace('\n', '')
    mp4_url= retrieve_mp4_data(slug)
    regex = re.compile('[^a-zA-Z0-9_]')
    output_path = (bpath +"/Clips/" + slug + ".mp4")    
    urllib.request.urlretrieve(mp4_url, output_path,reporthook=dl_progress)
    print("\nFinished "+slug +".mp4\n")
    ws['A'+str(temp)] = '=HYPERLINK("Clips/" & F'+ str(temp) + ' & ' + '".mp4",F' +str(temp)+')'
    temp+=1
print('Am terminat de descărcat totul.')
workbook.save(filename=bpath+"/"+"Excel.xlsx")
os.remove(bpath+"/Slug.txt")
